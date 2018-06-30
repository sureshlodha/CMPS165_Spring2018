import csv
import os
import re
import sys
import openpyxl
from reformat import id_map, output_file, output_fields

INPUT_DIR = 'microsoft_data'
OUTPUT_DIR = 'microsoft_output'
DATA_FILES = [
    {'year': 2013, 'fn': 'Microsoft LERR_2013_H1.xlsx', 'startRow': 6, 'gap': False},
    {'year': 2013, 'fn': 'Microsoft LERR_2013_H2.xlsx', 'startRow': 6, 'gap': True},
    {'year': 2014, 'fn': 'Microsoft LERR_2014_H1.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2014, 'fn': 'Microsoft LERR_2014_H2.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2015, 'fn': 'Microsoft LERR_2015_H1.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2015, 'fn': 'Microsoft LERR_2015_H2.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2016, 'fn': 'Microsoft LERR_2016_H1.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2016, 'fn': 'Microsoft LERR_2016_H2.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2017, 'fn': 'Microsoft LERR_2017_H1.xlsx', 'startRow': 8, 'gap': True},
    {'year': 2017, 'fn': 'Microsoft LERR_2017_H2.xlsx', 'startRow': 8, 'gap': True}
]

GAP_MAPPING = [
    (['A'], 'country', lambda name: name),
    (['C'], 'requests', lambda reqs: reqs),
    (['D'], 'accounts', lambda accounts: accounts),
    (['C', 'G', 'L'], 'accepted', lambda reqs, rej1, rej2: reqs - rej1 - rej2)
]

NO_GAP_MAPPING = [
    (['A'], 'country', lambda name: name),
    (['B'], 'requests', lambda reqs: reqs),
    (['C'], 'accounts', lambda accounts: accounts),
    (['B', 'I', 'K'], 'accepted', lambda reqs, rej1, rej2: reqs - rej1 - rej2)
]

RENAME = {
    'Bosnia And Herzegovina': 'Bosnia and Herzegovina',
    'Cote D\'Ivoire': 'Ivory Coast',
    'Korea': 'South Korea',
    'Russian Federation': 'Russia',
    'Korea, Republic Of': 'South Korea',
    'Urugay': 'Uruguay',
    'Turkey*': 'Turkey',
    'Venezuala':'Venezuela',
    'French Southern Territories': 'French Southern and Antarctic Lands'
}

# Set working directory to script directory
abspath = os.path.abspath(__file__)
dname = os.path.dirname(abspath)
os.chdir(dname)

# Create output directory
if not os.path.exists(OUTPUT_DIR):
    os.makedirs(OUTPUT_DIR)

# year	id	country	requests	accounts	percentAccepted
def load_data_file(file_info):
    wb = openpyxl.load_workbook('microsoft_data/' + file_info['fn'], data_only=True)
    ws = wb.active

    data = []
    row = file_info['startRow']
    mapping = GAP_MAPPING if file_info['gap'] else NO_GAP_MAPPING
    while True:
        if ws['A{}'.format(row)].value == None:
            break
        country = {}
        country['year'] = file_info['year']
        for cols, result_name, transform in mapping:
            values = list(map(lambda col: ws['{}{}'.format(col, row)].value or 0, cols))
            result = transform(*values)
            country[result_name] = result
        data.append(country)
        row += 1
    return data


def merge_country(c1, c2):
    result = {}
    for key in c1:
        if key in ['country', 'year'] :
            result[key] = c1[key]
        else:
            result[key] = c1[key] + c2[key]
    return result

def merge_year_halfs(d1, d2):
    results = []
    d2_map = {}

    for country in d2:
        d2_map[country['country']] = country

    for country in d1:
        name = country['country']
        if name in d2_map:
            results.append(merge_country(country, d2_map[name]))
            d2_map.pop(name)
        else:
            results.append(country)
    
    for name in d2_map:
        results.append(d2_map[name])

    return results


def add_ids(data, id_map):
    for country in data:
        name = country['country']
        if name in id_map:
            country['id'] = id_map[name]
        else:
            print('No id for', name)

all_data = []
for i in range(0, len(DATA_FILES) - 1, 2):
    half_one = load_data_file(DATA_FILES[i])
    half_two = load_data_file(DATA_FILES[i + 1])
    year_data = merge_year_halfs(half_one, half_two)

    for country in year_data:
        if country['country'] in RENAME:
            country['country'] = RENAME[country['country']]

        if country['requests'] != 0:
            country['percentAccepted'] = country['accepted'] / country['requests'] * 100
        else: 
            country['percentAccepted'] = 0
        country.pop('accepted')
        all_data.append(country)

add_ids(all_data, id_map)
output_file(all_data, 'microsoft_output/all_microsoft.tsv', output_fields)

# for col in ws.iter_cols(min_row=file_info['startRow'], max_col=12, max_row=50):
#    for cell in col:
#        print(cell)
